# File: excelsummary.py

import os
import platform
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
import time 
import re # เพิ่มการ import re

# ไลบรารีสำหรับควบคุม Excel COM (Windows เท่านั้น)
import win32com.client
import pythoncom

# นำเข้าคลาส Data จาก backend.py
# สมมติว่ามีไฟล์ backend.py ที่มีคลาส Data และเมธอด parse_amount และ format_thai_date
from Backend import Data

# ฟังก์ชันใหม่: ปิดเฉพาะแท็บไฟล์ Excel ที่เปิดอยู่ (ถ้ามี)
def close_excel_file_if_open(filename):
    pythoncom.CoInitialize()  # เรียก COM สำหรับ thread นี้
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        for wb in excel.Workbooks:
            # เปรียบเทียบชื่อไฟล์แบบไม่สนใจ case
            # ใช้ os.path.basename เพื่อให้แน่ใจว่าเปรียบเทียบแค่ชื่อไฟล์ ไม่รวมพาธเต็ม
            if os.path.basename(filename).lower() == os.path.basename(wb.FullName).lower():
                print(f"📄 พบไฟล์ {filename} ที่เปิดอยู่ใน Excel — กำลังปิดแท็บ")
                wb.Close(False)  # False = ปิดโดยไม่บันทึกซ้ำ
                return True
    except Exception as e:
        print("❌ ไม่สามารถตรวจสอบหรือปิดเอกสาร Excel:", e)
    return False

# ฟังก์ชันใหม่: บันทึกไฟล์ Excel พร้อม retry และปิดแท็บ Excel เฉพาะไฟล์นั้นถ้ายังเปิดอยู่
def save_excel_with_retry(wb, filename="Summary_Output.xlsx", max_retries=3):
    for attempt in range(max_retries):
        try:
            wb.save(filename)
            print(f"✅ {filename} created successfully!")
            if platform.system() == "Windows":
                os.startfile(filename)
            return True
        except PermissionError:
            print(f"⚠️ ไม่สามารถบันทึกไฟล์ {filename} ได้ อาจยังเปิดอยู่ใน Excel")
            print("🔄 กำลังพยายามปิดเฉพาะแท็บของไฟล์นั้น...")
            closed = close_excel_file_if_open(filename)
            if not closed:
                print("⏳ รอ 2 วินาทีแล้วลองใหม่...")
            time.sleep(2)
    print("❌ ไม่สามารถบันทึกไฟล์ได้ กรุณาปิดไฟล์ด้วยตนเองแล้วลองใหม่อีกครั้ง")
    return False


def create_excel_summary(data_list, transaction_info, filename="Summary_Output.xlsx"):
    """
    สร้างไฟล์ Excel สรุปยอดโดยใช้ข้อมูลที่รับมาและใส่สูตรคำนวณ
    - data_list: รายการพัสดุทั้งหมด
    - transaction_info: dict ข้อมูลสรุป (วันที่รับ, จ่ายให้)
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "สรุปยอด"

        # --- การตั้งค่า Font และ Border ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        thai_font = Font(name='TH Sarabun New', size=11)
        bold_thai_font = Font(name='TH Sarabun New', size=11, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        top_center_align = Alignment(horizontal='center', vertical='top')

        # --- ส่วนหัวตาราง ---
        # ปรับการ Merge Cells ให้ครอบคลุมแค่ A ถึง J
        ws.merge_cells('A1:J1')
        ws['A1'] = "เล่มที่.......... เลขที่..........."
        ws['A1'].font = bold_thai_font
        ws['A1'].alignment = right_align

        ws.merge_cells('A2:J2')
        ws['A2'] = "ชื่อ คณะเกษตรศาสตร์"
        ws['A2'].font = bold_thai_font
        ws['A2'].alignment = center_align

        ws.merge_cells('A3:J3')
        ws['A3'] = "งบประมาณรายได้ปี 2568 ภาควิชาอุตสาหกรรมการเกษตร คณะเกษตรศาสตร์ฯ"
        ws['A3'].font = bold_thai_font
        ws['A3'].alignment = center_align
        
        # --- Headers หลักและย่อย (ปรับตามที่ต้องการ) ---
        ws.merge_cells('A4:A5') # วันที่
        ws.merge_cells('B4:B5') # รายการ
        ws.merge_cells('C4:E4') # รับ
        ws.merge_cells('F4:H4') # จ่าย
        ws.merge_cells('I4:J4') # คงเหลือ (ปรับให้ครอบคลุม J4 เท่านั้น)
        
        # Main headers text
        ws['A4'] = 'วันที่'
        ws['B4'] = 'รายการ'
        ws['C4'] = 'รับ'
        ws['F4'] = 'จ่าย'
        ws['I4'] = 'คงเหลือ'

        # Sub-headers for 'รับ'
        ws['C5'] = 'ใบรับที่'
        ws['D5'] = 'จำนวน'
        ws['E5'] = 'บาท'

        # Sub-headers for 'จ่าย'
        ws['F5'] = 'ใบรับที่' # จะใช้เป็น "รวมจ่าย" แทน
        ws['G5'] = 'จำนวน'
        ws['H5'] = 'บาท'

        # Sub-headers for 'คงเหลือ'
        ws['I5'] = 'จำนวน'
        ws['J5'] = 'บาท'

        # Apply styles to headers (A4 to J5)
        for row_idx in range(4, 6):
            for col_idx in range(1, 11): # Columns A to J
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = bold_thai_font
                cell.alignment = center_align
                cell.border = thin_border
        
        # --- การตั้งค่าความกว้างคอลัมน์ (ปรับตามคอลัมน์ที่เหลือ) ---
        col_widths = {
            'A': 15, # วันที่
            'B': 45, # รายการ
            'C': 15, 'D': 10, 'E': 12, # รับ(ใบรับที่, จำนวน, บาท)
            'F': 15, 'G': 10, 'H': 12, # จ่าย(ใบรับที่, จำนวน, บาท)
            'I': 10, 'J': 12, # คงเหลือ (จำนวน, บาท)
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        current_row = 6 # เริ่มต้นที่แถว 6 สำหรับข้อมูล

        # --- จัดกลุ่มข้อมูลตาม (purchase_date, received_from, invoice_no) ---
        # ต้องเรียงลำดับข้อมูลก่อน เพื่อให้กลุ่มที่เหมือนกันอยู่ติดกันและแสดงผลตามลำดับวันที่
        data_list_sorted = sorted(data_list, key=lambda x: (x[7], x[5], x[6])) 
        
        grouped_data = defaultdict(list)
        data_handler_instance = Data() 
        for item in data_list_sorted: # ใช้อันที่เรียงแล้ว
            # ใช้ tuple (วันที่, บริษัท, ใบรับที่) เป็น key ในการจัดกลุ่ม
            group_key = (item[7], item[5], item[6]) 
            grouped_data[group_key].append(item)

        # ตัวแปรสำหรับเก็บยอดคงเหลือสะสม
        cumulative_balance_qty = 0
        cumulative_balance_baht = 0.0

        # วนลูปผ่านกลุ่มข้อมูลที่จัดใหม่ (ซึ่งตอนนี้ key คือ (วันที่, บริษัท, ใบรับที่))
        for group_key, items_in_group in grouped_data.items():
            purchase_date, company_name, invoice_no = group_key 

            # แถว "รับจาก"
            # A: วันที่ (ใช้ purchase_date จาก group_key)
            ws.cell(row=current_row, column=1, value=purchase_date).font = bold_thai_font
            ws.cell(row=current_row, column=1).alignment = center_align

            # B: รับจากบริษัท
            ws.cell(row=current_row, column=2, value=f"รับจาก {company_name}").font = bold_thai_font
            ws.cell(row=current_row, column=2).alignment = left_align
            
            # C: ใบรับที่ (ใช้ invoice_no จาก group_key)
            ws.cell(row=current_row, column=3, value=invoice_no).font = bold_thai_font
            ws.cell(row=current_row, column=3).alignment = center_align

            # D: จำนวน (จำนวนรายการในกลุ่มนี้) - เพิ่ม " รก."
            item_count_for_group = len(items_in_group) 
            ws.cell(row=current_row, column=4, value=f"{item_count_for_group} รก.").font = bold_thai_font
            ws.cell(row=current_row, column=4).alignment = center_align

            # E: บาท (รวมเงินของรายการในกลุ่มนี้) - คำนวณ จำนวน * ราคา
            total_price_for_group_E = 0
            for item in items_in_group:
                quantity_numeric = data_handler_instance.parse_amount(item[2]) 
                price_per_item = item[4] 
                total_price_for_group_E += (quantity_numeric * price_per_item)

            cell_group_total_price_E = ws.cell(row=current_row, column=5, value=total_price_for_group_E)
            cell_group_total_price_E.font = bold_thai_font
            cell_group_total_price_E.number_format = '#,##0.00'
            cell_group_total_price_E.alignment = right_align

            # F, G, H (จ่าย) - ใส่เครื่องหมาย "-"
            ws.cell(row=current_row, column=6, value="-").font = thai_font
            ws.cell(row=current_row, column=6).alignment = center_align
            ws.cell(row=current_row, column=7, value="-").font = thai_font
            ws.cell(row=current_row, column=7).alignment = center_align
            ws.cell(row=current_row, column=8, value="-").font = thai_font
            ws.cell(row=current_row, column=8).alignment = center_align


            # I: จำนวน (คงเหลือ) - ยอดสะสมจำนวน
            cumulative_balance_qty += item_count_for_group
            ws.cell(row=current_row, column=9, value=f"{cumulative_balance_qty} รก.").font = bold_thai_font
            ws.cell(row=current_row, column=9).alignment = center_align

            # J: บาท (คงเหลือ) - ยอดสะสมบาท
            cumulative_balance_baht += total_price_for_group_E
            cell_group_total_price_J = ws.cell(row=current_row, column=10, value=cumulative_balance_baht) 
            cell_group_total_price_J.font = bold_thai_font
            cell_group_total_price_J.number_format = '#,##0.00'
            cell_group_total_price_J.alignment = right_align

            # Apply border to the entire row
            for col_idx in range(1, 11): # Columns A to J
                ws.cell(row=current_row, column=col_idx).border = thin_border
            current_row += 1
            
            # รายละเอียดรายการย่อยภายใต้แต่ละกลุ่ม (invoice)
            for item in items_in_group:
                # [0:name, 1:category, 2:amount, 3:date_needed, 4:price, 5:received_from, 6:invoice_no, 7: purchase_date]
                item_name = item[0]
                amount_str = item[2] # เช่น "10 ด้าม"
                price_per_item = item[4] # เช่น 15.00
                
                # แยกตัวเลขและหน่วยนับออกจากกันอย่างถูกต้อง
                quantity_numeric = data_handler_instance.parse_amount(amount_str)
                # ใช้ regex เพื่อดึงส่วนที่เป็นตัวอักษรทั้งหมด (รวมสระและวรรณยุกต์)
                match = re.search(r'(\D+)$', amount_str.strip()) # หาตัวอักษรที่อยู่ท้ายสุดของสตริง
                unit = match.group(1).strip() if match else "" # ถ้าเจอให้ดึงมา ถ้าไม่เจอให้เป็นสตริงว่าง

                total_item_price = quantity_numeric * price_per_item

                # สร้างข้อความตาม format ที่ต้องการ
                if quantity_numeric > 1:
                    item_display_text = f"-{item_name} {quantity_numeric} {unit}@{price_per_item:.2f}.- = {total_item_price:.2f}.-"
                else:
                    item_display_text = f"-{item_name} {quantity_numeric} {unit}@{price_per_item:.2f}.-"


                # A: วันที่ - เว้นว่าง
                ws.cell(row=current_row, column=1, value="").font = thai_font
                ws.cell(row=current_row, column=1).alignment = center_align

                # B: รายการ (แสดงในรูปแบบใหม่)
                ws.cell(row=current_row, column=2, value=item_display_text).font = thai_font
                ws.cell(row=current_row, column=2).alignment = left_align

                # C, D, E (รับ) - เว้นว่าง
                ws.cell(row=current_row, column=3, value="").font = thai_font
                ws.cell(row=current_row, column=4, value="").font = thai_font
                ws.cell(row=current_row, column=5, value="").font = thai_font
                
                # F, G, H (จ่าย) - เว้นว่างสำหรับกรอกข้อมูล
                ws.cell(row=current_row, column=6, value="").font = thai_font
                ws.cell(row=current_row, column=7, value="").font = thai_font
                ws.cell(row=current_row, column=8, value="").font = thai_font

                # I, J (คงเหลือ) - เว้นว่าง
                ws.cell(row=current_row, column=9, value="").font = thai_font
                ws.cell(row=current_row, column=10, value="").font = thai_font

                for col_idx in range(1, 11): # Apply border to all cells in the row (A to J)
                    ws.cell(row=current_row, column=col_idx).border = thin_border
                
                current_row += 1

        # --- แถวสรุป "จ่ายให้" และ "รวมจ่าย" ---
        current_row += 1 # เว้น 1 บรรทัด
        
        # A: วันที่ (เว้นว่าง)
        ws.cell(row=current_row, column=1, value="").font = thai_font
        ws.cell(row=current_row, column=1).alignment = center_align

        # B: รายการ (จ่ายให้...)
        ws.cell(row=current_row, column=2, value=f"จ่ายให้ {transaction_info['paid_to']}").font = bold_thai_font
        ws.cell(row=current_row, column=2).alignment = left_align

        # C, D, E (รับ) - เว้นว่าง
        ws.cell(row=current_row, column=3, value="").font = thai_font
        ws.cell(row=current_row, column=4, value="").font = thai_font
        ws.cell(row=current_row, column=5, value="").font = thai_font

        # G: จำนวน (ใช้ค่าจาก cumulative_balance_qty ล่าสุด และเพิ่ม " รก." เข้าไปด้วย)
        ws.cell(row=current_row, column=7, value=f"{cumulative_balance_qty} รก.").font = bold_thai_font
        ws.cell(row=current_row, column=7).alignment = center_align

        # H: บาท (ใช้ค่าจาก cumulative_balance_baht ล่าสุด)
        ws.cell(row=current_row, column=8, value=cumulative_balance_baht).font = bold_thai_font
        ws.cell(row=current_row, column=8).number_format = '#,##0.00'
        ws.cell(row=current_row, column=8).alignment = right_align

        # I, J (คงเหลือ) - เว้นว่าง
        ws.cell(row=current_row, column=9, value="").font = thai_font
        ws.cell(row=current_row, column=10, value="").font = thai_font

        for col_idx in range(1, 11): # A to J
            ws.cell(row=current_row, column=col_idx).border = thin_border
        
        # --- Save the workbook ---
        return save_excel_with_retry(wb, filename)
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return False

def main():
    # สร้าง instance ของคลาส Data
    data_handler = Data()

    # --- ข้อมูลตัวอย่างที่มี บริษัทซ้ำ วันซ้ำ แต่คนละใบรับ ---

    # วันที่ 10/มิ.ย./68
    date_a = "10 มิ.ย.68"
    company_a = "บริษัท สยามพัฒนา จำกัด"
    company_b = "ร้านค้าส่งอุปกรณ์"

    # ใบรับที่ INV_A_001
    data_handler.appendlist("ปากกาเคมี", "วัสดุสำนักงาน", "5 ด้าม", date_a, 20.00, company_a, "INV_A_001", date_a)
    data_handler.appendlist("กระดาษโน้ต", "วัสดุสำนักงาน", "3 เล่ม", date_a, 15.00, company_a, "INV_A_001", date_a)
    
    # ใบรับที่ INV_A_002 (บริษัทเดียวกัน วันเดียวกัน แต่คนละใบรับ)
    data_handler.appendlist("แฟ้มเอกสาร", "วัสดุสำนักงาน", "10 อัน", date_a, 25.00, company_a, "INV_A_002", date_a)
    data_handler.appendlist("ยางลบ", "วัสดุสำนักงาน", "5 ก้อน", date_a, 5.00, company_a, "INV_A_002", date_a)

    # ใบรับที่ INV_B_001
    data_handler.appendlist("เทปใส", "วัสดุสำนักงาน", "2 ม้วน", date_a, 10.00, company_b, "INV_B_001", date_a)
    
    # ใบรับที่ INV_B_002 (บริษัทเดียวกัน วันเดียวกัน แต่คนละใบรับ)
    data_handler.appendlist("กาวแท่ง", "วัสดุสำนักงาน", "5 แท่ง", date_a, 8.00, company_b, "INV_B_002", date_a)
    data_handler.appendlist("ไม้บรรทัด", "วัสดุสำนักงาน", "3 อัน", date_a, 12.00, company_b, "INV_B_002", date_a)


    # วันที่ 15/มิ.ย./68
    date_b = "15 มิ.ย. 68"
    company_c = "บริษัท วัสดุก่อสร้าง"
    company_d = "บริษัท สยามพัฒนา จำกัด" # บริษัทซ้ำจาก date_a แต่คนละวัน คนละใบรับ

    data_handler.appendlist("หลอดไฟ LED", "วัสดุไฟฟ้า", "10 หลอด", date_b, 45.00, company_c, "INV_C_001", date_b)
    data_handler.appendlist("สายไฟ", "วัสดุไฟฟ้า", "1 ม้วน", date_b, 200.00, company_c, "INV_C_001", date_b)
    data_handler.appendlist("เม้าส์ไร้สาย", "อุปกรณ์คอมพิวเตอร์", "1 ชิ้น", date_b, 350.00, company_d, "INV_D_001", date_b)

    # วันที่ 20/มิ.ย./68
    date_c = "20 มิ.ย. 68"
    company_e = "ร้านค้าส่งอุปกรณ์" # บริษัทซ้ำจาก date_a แต่คนละวัน คนละใบรับ
    company_f = "บริษัท อาหารสด"

    data_handler.appendlist("ปากกาเจล", "วัสดุสำนักงาน", "12 ด้าม", date_c, 18.00, company_e, "INV_E_001", date_c)
    data_handler.appendlist("นมสด", "วัสดุบริโภค", "6 กล่อง", date_c, 30.00, company_f, "INV_F_001", date_c)

    data_handler.appendlist("เม้าส์ไร้สาย", "อุปกรณ์คอมพิวเตอร์", "1 ชิ้น", date_b, 350.00, company_d, "INV_D_099", date_b)

    # กำหนดข้อมูล transaction_info
    transaction_info = {
        "paid_to": "นาย ก. (ผู้รับของ)"
    }

    # เรียกใช้ฟังก์ชันสร้าง Excel
    create_excel_summary(data_handler.list, transaction_info)
    print("✅ การสร้างไฟล์ Excel เสร็จสมบูรณ์")

if __name__ == "__main__":
    main()